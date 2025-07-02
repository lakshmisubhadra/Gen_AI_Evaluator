import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
import threading
import time
import os
import importlib
import sys
from batch_processing.summary import add_summary_sheet
import requests

# Import your batch UI utility
from batch_processing.batch_ui import BatchUI

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

class GenAIEvaluatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("GenAI Evaluator")
        self.root.geometry("1100x800")
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        # Configure main grid layout
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)


        # Create frames
        self.left_frame = ctk.CTkFrame(self.root)
        self.left_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        self.right_frame = ctk.CTkFrame(self.root)
        self.right_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")

        self.bottom_frame = ctk.CTkFrame(self.root)
        self.bottom_frame.grid(row=1, column=0, columnspan=2, padx=20, pady=(0, 20), sticky="ew")

        # Configure frame grids
        self.left_frame.grid_columnconfigure(0, weight=1)
        self.left_frame.grid_rowconfigure((0, 1, 2), weight=1)

        self.right_frame.grid_columnconfigure(0, weight=1)
        self.right_frame.grid_rowconfigure(0, weight=3)  # Metrics
        self.right_frame.grid_rowconfigure(1, weight=4)  # Batch
        self.right_frame.grid_rowconfigure(2, weight=3)  # Results

        self.bottom_frame.grid_columnconfigure((0, 1, 2), weight=1)

        # Initialize variables and UI
        self.initialize_variables()
        self.setup_left_frame()
        self.setup_right_frame()
        self.setup_bottom_frame()

    def initialize_variables(self):
        self.selected_metric = tk.StringVar(value="")
        self.selected_metric.trace_add("write", self.toggle_expected_response)
        self.api_key = "" # Set API key to blank by default
        self.api_url = "https://api.deepseek.com/v1/chat/completions"
        self.evaluation_thread = None
        self.is_evaluating = False
        self.is_batch_processing = False 

        self.settings_visible = False
        self.settings_frame = None
        self.accept_criteria = {
            "Correctness": 90, "Relevancy": 90, "Hallucination": 90,
            "Completeness": 90, "Bias": 90, "Toxicity": 90, "Consistency": 90
        }
        self.selected_criteria_metric = tk.StringVar(value="Correctness")

        self.metric_definitions = {
            "Correctness": "Verifies if generated content is factually correct against trusted sources.",
            "Relevancy": "Evaluates alignment with user intent and context.",
            "Hallucination": "Measures fabrication of false information.",
            "Completeness": "Assesses thoroughness of responses.",
            "Bias": "Quantifies unfair or discriminatory outputs.",
            "Toxicity": "Measures harmful or inappropriate content.",
            "Consistency": "Evaluates output stability across contexts."
        }

    def toggle_expected_response(self, *args):
        if self.selected_metric.get() == "Correctness":
            self.expected_label.grid()
            self.expected_textbox.grid()
        else:
            self.expected_label.grid_remove()
            self.expected_textbox.grid_remove()

    def setup_left_frame(self):

        # Question to Chatbot
        question_label = ctk.CTkLabel(self.left_frame, text="Question to Chatbot", font=("Arial", 14, "bold"))
        question_label.grid(row=0, column=0, padx=20, pady=(20, 15), sticky="nw")

        self.question_textbox = ctk.CTkTextbox(self.left_frame, height=150, wrap="word")
        self.question_textbox.grid(row=0, column=0, padx=20, pady=(50, 20), sticky="nsew")

        # Chatbot Response
        response_label = ctk.CTkLabel(self.left_frame, text="Chatbot Response", font=("Arial", 14, "bold"))
        response_label.grid(row=1, column=0, padx=20, pady=(20, 15), sticky="nw")

        self.response_textbox = ctk.CTkTextbox(self.left_frame, height=150, wrap="word")
        self.response_textbox.grid(row=1, column=0, padx=20, pady=(50, 20), sticky="nsew")

        # Expected Response
        self.expected_label = ctk.CTkLabel(self.left_frame, text="Expected Response", font=("Arial", 14, "bold"))
        self.expected_label.grid(row=2, column=0, padx=20, pady=(20, 15), sticky="nw")
        self.expected_label.grid_remove()

        self.expected_textbox = ctk.CTkTextbox(self.left_frame, height=150, wrap="word")
        self.expected_textbox.grid(row=2, column=0, padx=20, pady=(50, 20), sticky="nsew")
        self.expected_textbox.grid_remove()

    def setup_right_frame(self):
        # Metrics Section
        self.setup_metrics_section()

        # Batch Processing Section (use BatchUI)
        self.setup_batch_section()

        # Results Section
        self.setup_results_section()

    def setup_metrics_section(self):
        metrics_frame = ctk.CTkFrame(self.right_frame)
        metrics_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        header_frame = ctk.CTkFrame(metrics_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(10, 15), padx=20)

        metrics_label = ctk.CTkLabel(header_frame, text="Evaluation Metrics", font=("Arial", 16, "bold"))
        metrics_label.pack(side="left")

        info_icon = ctk.CTkButton(header_frame, text="ⓘ", width=30, height=30,
                                  command=self.show_metrics_info,
                                  fg_color="transparent", hover_color="#5a6268",text_color="black")  # Add this line
        info_icon.pack(side="left", padx=(10, 0))

        settings_icon = ctk.CTkButton(header_frame, text="🔧", width=30, height=30,
                                      command=self.toggle_settings,
                                      fg_color="#6c757d", hover_color="#5a6268")
        settings_icon.pack(side="right")

        metrics_grid_frame = ctk.CTkFrame(metrics_frame, fg_color="transparent")
        metrics_grid_frame.pack(fill="both", expand=True, padx=20, pady=10)

        metrics = ["Correctness", "Relevancy", "Hallucination", "Completeness",
                   "Bias", "Toxicity", "Consistency"]
        
        for i, metric in enumerate(metrics):
            row = i // 3
            col = i % 3
            radio_btn = ctk.CTkRadioButton(metrics_grid_frame, text=metric,
                                           variable=self.selected_metric, value=metric)
            radio_btn.grid(row=row, column=col, padx=10, pady=10, sticky="w")

    def setup_batch_section(self):
        # Use BatchUI for all batch processing UI and logic
        self.batch_ui_frame = ctk.CTkFrame(self.right_frame)
        self.batch_ui_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        self.batch_ui = BatchUI(
            parent_frame=self.batch_ui_frame,
            api_key=self.api_key,
            api_url=self.api_url,
            accept_criteria=self.accept_criteria
        )

    def setup_results_section(self):
        results_frame = ctk.CTkFrame(self.right_frame)
        results_frame.grid(row=2, column=0, padx=5, pady=5, sticky="nsew")

        results_label = ctk.CTkLabel(results_frame, text="Evaluation Results",
                                     font=("Arial", 16, "bold"))
        results_label.pack(pady=(10, 20), padx=20, anchor="nw")

        results_info_frame = ctk.CTkFrame(results_frame, fg_color="transparent")
        results_info_frame.pack(fill="x", padx=20, pady=5)

        score_frame = ctk.CTkFrame(results_info_frame, fg_color="transparent")
        score_frame.pack(side="left", padx=(0, 20))
        ctk.CTkLabel(score_frame, text="Score:", font=("Arial", 14)).pack(side="left")
        self.score_value = ctk.CTkLabel(score_frame, text="0%", font=("Arial", 14, "bold"))
        self.score_value.pack(side="left", padx=10)

        time_frame = ctk.CTkFrame(results_info_frame, fg_color="transparent")
        time_frame.pack(side="left", padx=(0, 20))
        ctk.CTkLabel(time_frame, text="Time:", font=("Arial", 14)).pack(side="left")
        self.time_value = ctk.CTkLabel(time_frame, text="0h 0m 0s", font=("Arial", 14, "bold"))
        self.time_value.pack(side="left", padx=10)

        status_frame = ctk.CTkFrame(results_info_frame, fg_color="transparent")
        status_frame.pack(side="left")
        ctk.CTkLabel(status_frame, text="Status:", font=("Arial", 14)).pack(side="left")
        self.status_value = ctk.CTkLabel(status_frame, text="N/A", font=("Arial", 14, "bold"))
        self.status_value.pack(side="left", padx=10)

        self.reason_textbox = ctk.CTkTextbox(results_frame, height=100, state="disabled", wrap="word")
        self.reason_textbox.pack(fill="both", expand=True, padx=20, pady=(0, 20))


    def setup_bottom_frame(self):

        self.start_button = ctk.CTkButton(self.bottom_frame, text="Start",
                                          command=self.start_evaluation,
                                          fg_color="#28a745", hover_color="#218838")
        self.start_button.grid(row=0, column=0, padx=20, pady=10)

        self.stop_button = ctk.CTkButton(self.bottom_frame, text="Stop",
                                         command=self.stop_evaluation,
                                         fg_color="#dc3545", hover_color="#c82333", state="disabled")
        self.stop_button.grid(row=0, column=1, padx=20, pady=10)

        self.clear_button = ctk.CTkButton(self.bottom_frame, text="Clear",
                                          command=self.clear_fields,
                                          fg_color="#6c757d", hover_color="#5a6268")
        self.clear_button.grid(row=0, column=2, padx=20, pady=10)





        #UNTILL HERE

    def show_metrics_info(self):
        info_window = ctk.CTkToplevel(self.root)
        info_window.title("Evaluation Metrics Information")
        info_window.geometry("600x400")
        info_window.grab_set()  # Make the window modal
        
        # Add a scrollable frame
        scroll_frame = ctk.CTkScrollableFrame(info_window)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Add title
        title_label = ctk.CTkLabel(scroll_frame, text="Evaluation Metrics Definitions", 
                                  font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 20), anchor="w")
        
        # Add definitions
        for metric, definition in self.metric_definitions.items():
            metric_frame = ctk.CTkFrame(scroll_frame)
            metric_frame.pack(fill="x", pady=10)
            
            metric_label = ctk.CTkLabel(metric_frame, text=f"{metric}:", 
                                       font=("Arial", 14, "bold"))
            metric_label.pack(anchor="w", padx=10, pady=(10, 5))
            
            definition_label = ctk.CTkLabel(metric_frame, text=definition, 
                                          font=("Arial", 12), wraplength=540, 
                                          justify="left")
            definition_label.pack(anchor="w", padx=10, pady=(0, 10))
        
        # Add close button
        close_button = ctk.CTkButton(info_window, text="Close", 
                                    command=info_window.destroy,
                                    fg_color="#6c757d", hover_color="#5a6268")
        close_button.pack(pady=20)
    
    def toggle_settings(self):
        if self.settings_visible:
            self.hide_settings()
        else:
            self.show_settings()
    
    def show_settings(self):
        # Create settings sidebar
        self.settings_frame = ctk.CTkFrame(self.root, width=300)
        self.settings_frame.place(relx=1, rely=0, relheight=1, anchor="ne")
        
        # Add title
        settings_title = ctk.CTkLabel(self.settings_frame, text="Settings", 
                                     font=("Arial", 16, "bold"))
        settings_title.pack(pady=(20, 30), padx=20)

        # --- API Key Section ---
        apikey_label = ctk.CTkLabel(self.settings_frame, text="DeepSeek API Key", font=("Arial", 14, "bold"))
        apikey_label.pack(anchor="w", padx=20, pady=(0, 5))

        # --- Add this: API Key Entry ---
        self.apikey_entry = ctk.CTkEntry(self.settings_frame, width=260)
        self.apikey_entry.pack(padx=20, pady=(0, 10))
        self.apikey_entry.insert(0, self.api_key)# will be set to blank by default
        self.apikey_entry.bind("<FocusOut>", self.update_api_key)
        self.apikey_entry.bind("<Return>", self.update_api_key)
        
        # Accept criteria section
        criteria_label = ctk.CTkLabel(self.settings_frame, text="Accept Criteria", 
                                     font=("Arial", 14, "bold"))
        criteria_label.pack(anchor="w", padx=20, pady=(0, 10))
        
        # Metric selection dropdown
        metric_selection_frame = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        metric_selection_frame.pack(fill="x", padx=20, pady=10)
        
        metric_label = ctk.CTkLabel(metric_selection_frame, text="Metric:", 
                                   font=("Arial", 12))
        metric_label.pack(side="left")
        
        metrics = ["Correctness", "Relevancy", "Hallucination", "Completeness", 
                  "Bias", "Toxicity", "Consistency"]
        self.metric_dropdown = ctk.CTkOptionMenu(metric_selection_frame, 
                                               values=metrics,
                                               variable=self.selected_criteria_metric,
                                               command=self.update_threshold_display)
        self.metric_dropdown.pack(side="left", padx=(10, 0), fill="x", expand=True)
        
        # Threshold slider
        threshold_frame = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        threshold_frame.pack(fill="x", padx=20, pady=10)
        
        threshold_label = ctk.CTkLabel(threshold_frame, text="Threshold:", 
                                      font=("Arial", 12))
        threshold_label.pack(side="left")
        
        self.threshold_value_label = ctk.CTkLabel(threshold_frame, 
                                                text=f"{self.accept_criteria[self.selected_criteria_metric.get()]}%", 
                                                font=("Arial", 12, "bold"))
        self.threshold_value_label.pack(side="right")
        
        self.threshold_slider = ctk.CTkSlider(self.settings_frame, 
                                             from_=1, to=100, 
                                             number_of_steps=99,
                                             command=self.update_threshold_value)
        self.threshold_slider.pack(fill="x", padx=20, pady=10)
        self.threshold_slider.set(self.accept_criteria[self.selected_criteria_metric.get()])
        
        # Done button
        done_button = ctk.CTkButton(self.settings_frame, text="Done", 
                                   command=self.hide_settings,
                                   fg_color="#28a745", hover_color="#218838")
        done_button.pack(pady=30)
        
        # Animation to slide in
        settings_width = 300
        screen_width = self.root.winfo_width()
        slide_steps = 15
        step_size = settings_width / slide_steps
        
        for i in range(slide_steps + 1):
            position = 1 - (i * step_size / screen_width)
            self.settings_frame.place(relx=position, rely=0, relheight=1, anchor="ne")
            self.root.update()
            time.sleep(0.01)
        
        self.settings_visible = True

    def update_api_key(self, event=None):
        new_key = self.apikey_entry.get().strip()
        
        self.api_key = new_key # Always update, even if blank
        print(f"API key updated to: {self.api_key}")
    
    def hide_settings(self):
        if self.settings_frame:
            self.update_api_key()
            # Animation to slide out
            settings_width = 300
            screen_width = self.root.winfo_width()
            slide_steps = 15
            step_size = settings_width / slide_steps
            
            for i in range(slide_steps + 1):
                position = 1 - (settings_width / screen_width) + (i * step_size / screen_width)
                self.settings_frame.place(relx=position, rely=0, relheight=1, anchor="ne")
                self.root.update()
                time.sleep(0.01)
            
            self.settings_frame.destroy()
            self.settings_frame = None
            self.settings_visible = False
    
    def update_threshold_display(self, _=None):  # Changed parameter name to _ to indicate it's unused
        selected_metric = self.selected_criteria_metric.get()
        current_value = self.accept_criteria.get(selected_metric, 90)
        self.threshold_slider.set(current_value)
        self.threshold_value_label.configure(text=f"{current_value}%")
    
    def update_threshold_value(self, value):
        value = int(value)
        selected_metric = self.selected_criteria_metric.get()
        self.accept_criteria[selected_metric] = value
        self.threshold_value_label.configure(text=f"{value}%")
    
    def validate_inputs(self):
            # Check if API key is provided
        if not self.api_key.strip():
            messagebox.showerror("Input Error", "Input DeepSeek API key")
            return False

        # Check if text boxes are filled
        question = self.question_textbox.get("1.0", "end-1c").strip()
        response = self.response_textbox.get("1.0", "end-1c").strip()
        
        if not question or not response:
            messagebox.showerror("Input Error", "Please fill the Question and Chatbot Response fields before starting evaluation.")
            return False
        
        # For Correctness metric, expected response is required
        if self.selected_metric.get() == "Correctness":
            expected = self.expected_textbox.get("1.0", "end-1c").strip()
            if not expected:
                messagebox.showerror("Input Error", "Expected Response is required for Correctness evaluation.")
                return False
        
        # Check if a metric is selected
        if not self.selected_metric.get():
            messagebox.showerror("Selection Error", "Please select an evaluation metric before starting.")
            return False
        
        return True
    
    def start_evaluation(self):
        self.stop_requested = False
        if not self.validate_inputs():
            return
        
        # Get inputs
        question = self.question_textbox.get("1.0", "end-1c").strip()
        response = self.response_textbox.get("1.0", "end-1c").strip()
        expected = self.expected_textbox.get("1.0", "end-1c").strip()
        metric = self.selected_metric.get()
        
        # Update UI state
        self.is_evaluating = True
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self.clear_button.configure(state="disabled")
        self.root.config(cursor="wait")
        
        # Show evaluation in progress
        self.score_value.configure(text="Evaluating...")
        self.time_value.configure(text="0h 0m 0s")
        self.status_value.configure(text="In Progress")
        
        self.reason_textbox.configure(state="normal")
        self.reason_textbox.delete("1.0", "end")
        self.reason_textbox.insert("1.0", "Evaluation in progress, please wait...")
        self.reason_textbox.configure(state="disabled")
        
        # Start evaluation in a separate thread
        self.evaluation_thread = threading.Thread(
            target=self.run_evaluation_thread,
            args=(question, response, expected, metric)
        )
        self.evaluation_thread.daemon = True
        self.evaluation_thread.start()
    def run_evaluation_thread(self, question, response, expected, metric):
        start_time = time.time()
        status = "N/A"
        try:
            module_name = f"scoring_files.{metric.lower()}"
            scoring_module = importlib.import_module(module_name)
            evaluation_function = getattr(scoring_module, f"evaluate_{metric.lower()}")

            try:
                    # Try online evaluation
                if metric == "Correctness":
                     result, _ = evaluation_function(
                    question, response, expected, self.api_key, self.api_url, stop_requested=lambda: self.stop_requested
                        )
                else:
                        result, _ = evaluation_function(
                            question, response, self.api_key, self.api_url
                        )
            except (requests.RequestException, ConnectionError, TimeoutError, OSError) as api_exc:
                # API/network failure: switch to offline mode
                self.root.after(0, lambda: self.reason_textbox.configure(state="normal"))
                self.root.after(0, lambda: self.reason_textbox.delete("1.0", "end"))
                self.root.after(0, lambda: self.reason_textbox.insert("1.0", "Offline mode: Using cached models"))
                self.root.after(0, lambda: self.reason_textbox.configure(state="disabled"))
                # Call a fallback scoring function (implement this in your scoring module)
                if hasattr(scoring_module, "evaluate_offline"):
                    if metric == "Correctness":
                        result, _ = scoring_module.evaluate_offline(
                            question, response, expected
                        )
                    else:
                        result, _ = scoring_module.evaluate_offline(
                            question, response
                        )
                else:
                    # Provide a default basic scoring if offline scoring is not implemented
                    result = {"score": 50, "reason": "Basic offline scoring used due to network error."}

            if self.stop_requested:
                self.root.after(0, self.reset_ui_after_evaluation)
                return

            elapsed_time = time.time() - start_time
            hours, remainder = divmod(elapsed_time, 3600)
            minutes, seconds = divmod(remainder, 60)
            time_str = f"{int(hours)}h {int(minutes)}m {int(seconds)}s"

            score = result.get("score", 0)
            threshold = self.accept_criteria[metric]
            if metric in ["Toxicity", "Bias", "Hallucination"]:
                status = "Failed" if score >= threshold else "Passed"
            else:
                status = "Passed" if score >= threshold else "Failed"

            self.root.after(0, lambda: self.update_results(result, time_str, status))
        except Exception as e:
            error_msg = str(e)
            self.root.after(0, lambda: messagebox.showerror("API Error", f"Error during evaluation: {error_msg}"))
            self.root.after(0, lambda: self.reset_ui_after_evaluation())
    def reset_ui_after_evaluation(self):
        self.is_evaluating = False
        self.start_button.configure(state="normal")
        self.stop_button.configure(state="disabled")
        self.clear_button.configure(state="normal")
        self.root.config(cursor="")
    
    def update_results(self, result, time_str, status):
        score = result.get("score", 0)
        reason = result.get("reason", "No reason provided.")
        
        self.score_value.configure(text=f"{score}%")
        self.time_value.configure(text=time_str)
        
        # Set status with color
        self.status_value.configure(
            text=status,
            text_color="#28a745" if status == "Passed" else "#dc3545"
        )
        
        self.reason_textbox.configure(state="normal")
        self.reason_textbox.delete("1.0", "end")
        self.reason_textbox.insert("1.0", reason)
        self.reason_textbox.configure(state="disabled")
        
        self.reset_ui_after_evaluation()
    
    def stop_evaluation(self):
        # Set the stop flag
        self.stop_requested = True

        if self.is_evaluating:
            messagebox.showinfo("Stop", "Evaluation Stopped")
            self.reset_ui_after_evaluation()
        if self.is_batch_processing:
            messagebox.showinfo("Stop", "Stopping batch processing. Please wait for the current row to finish.")
            # UI will be reset in _process_batch when the loop sees the stop flag

    def update_batch_ui_stopped(self):
        self.is_batch_processing = False
        self.batch_button.configure(state="normal")
        self.upload_button.configure(state="normal")
        self.download_button.configure(state="disabled")
        self.status_label.configure(
            text="Batch processing stopped by user.",
            text_color="#dc3545"
        )
        self.progress_bar.set(0)
    
    def clear_fields(self):
        # Clear all input fields and results for single evaluation
        self.question_textbox.delete("1.0", "end")
        self.response_textbox.delete("1.0", "end")
        self.expected_textbox.delete("1.0", "end")
        
        self.selected_metric.set("")
        
        self.score_value.configure(text="0%")
        self.time_value.configure(text="0h 0m 0s")
        self.status_value.configure(text="N/A", text_color=("gray10", "gray90"))
        
        self.reason_textbox.configure(state="normal")
        self.reason_textbox.delete("1.0", "end")
        self.reason_textbox.configure(state="disabled")
        
        # Reset batch UI using BatchUI's method
        if hasattr(self, "batch_ui"):
            self.batch_ui.clear_fields()

        self.is_batch_processing = False
        self.stop_requested = False
    
    def upload_excel(self):
        """Handle Excel file upload."""
        from tkinter import filedialog
        import pandas as pd

        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if not file_path:
            return

        try:
            # Read the Excel file to validate
            df = pd.read_excel(file_path)
            # Normalize column names for validation

            df.columns = [col.strip().lower() for col in df.columns]
            required_columns = ["question to chatbot", "chatbot response", "expected response"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messagebox.showerror("Invalid Excel File",
                                 f"Missing required columns: {', '.join(missing_columns)}")
                return
        except Exception as e:
            messagebox.showerror("Error", f"Error reading Excel file: {str(e)}")
            return
        
        #Update UI with selected file
        self.uploaded_file_path = file_path
        filename = os.path.basename(file_path)
        self.file_label.configure(text=f"Selected: {filename}", text_color="white")
        self.batch_button.configure(state="normal")
        self.download_button.configure(state="disabled")
        self.processed_file_path = None


    
    def run_batch_evaluation(self):
        self.stop_requested = False
        """Run batch evaluation on the uploaded Excel file."""
        if not self.uploaded_file_path or self.is_batch_processing:
            return
        
        self.is_batch_processing = True
        self.batch_button.configure(state="disabled")
        self.upload_button.configure(state="disabled")
        self.download_button.configure(state="disabled")
        self.progress_bar.set(0)
        self.status_label.configure(text="Starting batch processing...")
        
        # Start processing in a separate thread
        self.batch_thread = threading.Thread(
            target=self._process_batch,
            daemon=True
        )
        self.batch_thread.start()
 
# pyright: ignore[reportUnusedFunction]
    def _process_batch(self):
        """Process the batch in a separate thread."""
    # ... method code ...
        try:
            import pandas as pd
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            
            start_time = time.time()
            df = pd.read_excel(self.uploaded_file_path)
            # Normalize column names
            df.columns = [col.strip().lower() for col in df.columns]
            required_columns = ["question to chatbot", "chatbot response", "expected response"]
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
            # Initialize new columns for results
            df["Question to chatbot"] = df["question to chatbot"]
            df["Chatbot Response"] = df["chatbot response"]
            df["Expected Response"] = df["expected response"]
            df.drop(columns=required_columns, inplace=True, errors='ignore')

            # Initialize progress bar
            self.root.after(0, lambda: self.progress_bar.set(0))
            self.root.after(0, lambda: self.status_label.configure(text="Processing batch..."))
            # Get total number of rows for progress tracking
            importlib.reload(importlib.import_module("scoring_files.correctness"))
            importlib.reload(importlib.import_module("scoring_files.relevancy"))
            importlib.reload(importlib.import_module("scoring_files.hallucination"))
            importlib.reload(importlib.import_module("scoring_files.completeness"))
            importlib.reload(importlib.import_module("scoring_files.bias"))
            importlib.reload(importlib.import_module("scoring_files.toxicity"))
            importlib.reload(importlib.import_module("scoring_files.consistency"))
            importlib.reload(importlib.import_module("batch_processing.summary"))   
            total_rows = len(df)
            
            metrics = ["Correctness", "Relevancy", "Hallucination", "Completeness", 
                      "Bias", "Toxicity", "Consistency"]
            
            for metric in metrics:
                df[f"{metric} Score"] = None
                df[f"{metric} Status"] = None
                df[f"{metric} Reason"] = None

            # Process each row
            for index, row in df.iterrows():

                if self.stop_requested:
                    break  # Stop batch processing if stop requested


                # Update progress
                progress = (index + 1) / total_rows
                self.root.after(0, lambda: self.progress_bar.set(progress))
                self.root.after(0, lambda: self.status_label.configure(
                    text=f"Processing row {index+1}/{total_rows} ({progress*100:.1f}%)"
                ))

                # Process each metric
                for metric in metrics:
                    if self.stop_requested:
                        break

                    try:
                        # Get evaluation results
                        module_name = f"scoring_files.{metric.lower()}"
                        scoring_module = importlib.import_module(module_name)
                        evaluation_function = getattr(scoring_module, f"evaluate_{metric.lower()}")

                        if metric == "Correctness":
                            result, _ = evaluation_function(
                                row["Question to chatbot"],
                                row["Chatbot Response"],
                                row["Expected Response"],
                                self.api_key,
                                self.api_url
                            )
                        else:
                            result, _ = evaluation_function(
                                row["Question to chatbot"],
                                row["Chatbot Response"],
                                self.api_key,
                                self.api_url
                            )

                        score = result.get("score", 0)
                        threshold = self.accept_criteria.get(metric, 90)
                        if metric in ["Toxicity", "Bias", "Hallucination"]:
                            status = "Failed" if score >= threshold else "Passed"
                            print("bias status", status)
                        else:
                             status = "Passed" if score >= threshold else "Failed"

                        # Update DataFrame
                        df.at[index, f"{metric} Score"] = f"{score}%"
                        df.at[index, f"{metric} Status"] = status
                        df.at[index, f"{metric} Reason"] = result.get("reason", "")

                    except Exception as e:
                        print(f"Error processing {metric} for row {index}: {str(e)}")

            if self.stop_requested:
                self.root.after(0, lambda: self.update_batch_ui_stopped())
                return

            # Save with proper formatting
            output_dir = os.path.dirname(self.uploaded_file_path)
            base_name = os.path.splitext(os.path.basename(self.uploaded_file_path))[0]
            self.processed_file_path = os.path.join(output_dir, f"{base_name}_results.xlsx")

            # Create Excel writer object
            with pd.ExcelWriter(self.processed_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Results')

                # Add summary sheet
                add_summary_sheet(writer, df)

                # Get the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['Results']

                # Define styles
                header_font = Font(bold=True)
                center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                wrap_alignment = Alignment(wrap_text=True, vertical='top')

                # Apply formatting to all columns
                for col_idx, column in enumerate(df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[col_letter].width = 25

                    # Apply formatting to each cell in column
                    for row_idx, cell in enumerate(worksheet[col_letter], 1):
                        if row_idx == 1:
                            cell.font = header_font
                            cell.alignment = center_alignment
                        else:
                            if "Score" in column or "Status" in column:
                                cell.alignment = center_alignment
                            else:
                                cell.alignment = wrap_alignment

                # Auto-adjust row heights
                for row in worksheet.iter_rows():
                    for cell in row:
                        if "Reason" in cell.column_letter or cell.column_letter in ['A', 'B', 'C']:
                            worksheet.row_dimensions[cell.row].height = 60

            elapsed_time = time.time() - start_time
            hours, remainder = divmod(elapsed_time, 3600)
            minutes, seconds = divmod(remainder, 60)
            time_str = f"{int(hours)}h {int(minutes)}m {int(seconds)}s"

            self.root.after(0, lambda: self.update_batch_ui_completed(True, time_str))

        except Exception as e:
            self.root.after(0, lambda err=str(e): self.update_batch_ui_error(err))
    
    def update_batch_ui_completed(self, success, elapsed_time):
        """Update UI when batch processing completes."""
        self.is_batch_processing = False
        self.batch_button.configure(state="normal")
        self.upload_button.configure(state="normal")
        
        if success:
            self.download_button.configure(state="normal")
            self.status_label.configure(
                text=f"Batch processing completed in {elapsed_time}",
                text_color="#28a745"
            )
            self.progress_bar.set(1)
            messagebox.showinfo("Batch Processing Complete", 
                               f"Batch processing completed in {elapsed_time}")
        else:
            self.status_label.configure(
                text="Error saving results",
                text_color="#dc3545"
            )
    
    def update_batch_ui_error(self, error_message):
        """Update UI when batch processing encounters an error."""
        self.is_batch_processing = False
        self.batch_button.configure(state="normal")
        self.upload_button.configure(state="normal")
        self.status_label.configure(
            text=f"Error: {error_message}",
            text_color="#dc3545"
        )
        messagebox.showerror("Batch Processing Error", f"Error: {error_message}")
    
    def download_results(self):
        """Handle downloading the processed results."""
        if not self.processed_file_path:
            return
        
        from tkinter import filedialog
        
        save_path = filedialog.asksaveasfilename(
            title="Save Results",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=os.path.basename(self.processed_file_path)
        )
        
        if not save_path:
            return
        
        try:
            # Copy the processed file to the selected location
            import shutil
            shutil.copy2(self.processed_file_path, save_path)
            
            self.status_label.configure(
                text=f"Results saved to {os.path.basename(save_path)}",
                text_color="#28a745"
            )
            messagebox.showinfo("Download Complete", 
                               f"Results saved to {os.path.basename(save_path)}")
        except Exception as e:
            messagebox.showerror("Download Error", f"Error saving file: {str(e)}")

def main():
    root = ctk.CTk()
    app = GenAIEvaluatorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()