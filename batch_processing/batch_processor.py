import pandas as pd
import importlib
import time
import os
import sys
import random
import requests
from pathlib import Path
from requests.exceptions import HTTPError
from openpyxl.styles import Alignment

# Add parent directory to path to import scoring modules
sys.path.append(str(Path(__file__).parent.parent))

class RateLimiter:
    """Handles rate limiting with exponential backoff and jitter."""
    def __init__(self, max_retries=5, base_delay=1.0, max_delay=60.0):
        self.max_retries = max_retries
        self.base_delay = base_delay
        self.max_delay = max_delay
    
    def wait_and_retry(self, attempt):
        if attempt >= self.max_retries:
            raise Exception("Max retries exceeded")
        
        # Exponential backoff with jitter
        delay = min(self.base_delay * (2 ** attempt) + random.uniform(0, 1), self.max_delay)
        time.sleep(delay)
        return delay

class BatchProcessor:
    """
    Handles batch processing of GenAI evaluation tasks from Excel files with robust rate limiting.
    """

    def __init__(self, api_key, api_url, accept_criteria, request_delay=0.2):
        """
        Initialize the batch processor.

        Parameters:
        - api_key (str): The API key for DeepSeek
        - api_url (str): The API URL for DeepSeek
        - accept_criteria (dict): Dictionary of acceptance thresholds for each metric
        - request_delay (float): Delay in seconds between each metric evaluation to avoid rate limits
        """
        self.api_key = api_key
        self.api_url = api_url
        self.accept_criteria = accept_criteria
        self.request_delay = request_delay
        self.metrics = [
            "Correctness", "Relevancy", "Hallucination", "Completeness",
            "Bias", "Toxicity", "Consistency"
        ]
        self.required_columns = [
            "Question to chatbot", "Chatbot Response", "Expected Response"
        ]

    def validate_excel(self, file_path):
        """
        Validates that the Excel file contains the required columns.

        Parameters:
        - file_path (str): Path to the Excel file

        Returns:
        - tuple: (is_valid, error_message)
        """
        try:
            df = pd.read_excel(file_path)
            # Normalize column names for case insensitivity
            normalized_columns = [col.strip().lower() for col in df.columns]
            missing_columns = [col for col in self.required_columns 
                              if col.lower() not in normalized_columns]
            
            if missing_columns:
                return False, f"Missing required columns: {', '.join(missing_columns)}"
            return True, ""
        except Exception as e:
            return False, f"Error validating Excel file: {str(e)}"

    def process_batch(self, file_path, progress_callback=None, status_callback=None, stop_flag=None):
        """
        Process all rows in the Excel file for all metrics with rate limiting.

        Parameters:
        - file_path (str): Path to the Excel file
        - progress_callback (function): Optional callback function to report progress
        - status_callback (function): Optional callback for status messages
        - stop_flag (callable): Optional function that returns True if processing should stop

        Returns:
        - tuple: (processed_df, elapsed_time)
        """
        start_time = time.time()
        df = pd.read_excel(file_path)
        total_rows = len(df)
        
        # Preload scoring modules
        scoring_modules = {}
        for metric in self.metrics:
            module_name = f"scoring_files.{metric.lower()}"
            try:
                scoring_modules[metric] = importlib.import_module(module_name)
                # Reload to ensure fresh state
                importlib.reload(scoring_modules[metric])
            except ImportError as e:
                if status_callback:
                    status_callback(f"Error loading {metric} module: {str(e)}")
                scoring_modules[metric] = None

        # Create result columns for each metric
        for metric in self.metrics:
            df[f"{metric} Score"] = None
            df[f"{metric} Status"] = None
            df[f"{metric} Reason"] = None

        # Initialize rate limiter
        rate_limiter = RateLimiter(max_retries=5, base_delay=10.0, max_delay=120.0)

        # Process each row
        for idx, row in df.iterrows():
            if stop_flag and stop_flag():
                break

            # Get normalized column values (case-insensitive)
            question = self.get_row_value(row, "Question to chatbot")
            chatbot_response = self.get_row_value(row, "Chatbot Response")
            expected_response = self.get_row_value(row, "Expected Response")

            # Update progress if callback provided
            if progress_callback:
                progress_callback(idx, total_rows)

            # Process each metric for this row
            for metric in self.metrics:
                if stop_flag and stop_flag():
                    break
                
                # Add request delay between metrics
                time.sleep(self.request_delay)
                
                # Try evaluation with rate limiting
                for attempt in range(rate_limiter.max_retries):  # Max 5 attempts
                    try:
                        if scoring_modules.get(metric) is None:
                            df.at[idx, f"{metric} Reason"] = "Scoring module not available"
                            break
                            
                        evaluation_function = getattr(scoring_modules[metric], f"evaluate_{metric.lower()}")
                        
                        if metric == "Correctness":
                            result, _ = evaluation_function(
                                question, chatbot_response, expected_response,
                                self.api_key, self.api_url
                            )
                        else:
                            result, _ = evaluation_function(
                                question, chatbot_response,
                                self.api_key, self.api_url
                            )

                        score = result.get("score", 0)
                        threshold = self.accept_criteria.get(metric, 90)
                        
                        # Determine status based on metric type
                        if metric in ["Toxicity", "Bias", "Hallucination"]:
                            status = "Failed" if score >= threshold else "Passed"
                        else:
                            status = "Passed" if score >= threshold else "Failed"
                            
                        # Update DataFrame
                        df.at[idx, f"{metric} Score"] = f"{score}%"
                        df.at[idx, f"{metric} Status"] = status
                        df.at[idx, f"{metric} Reason"] = result.get("reason", "")
                        break  # Success, exit retry loop

                    except requests.exceptions.RequestException as e:
                        # Check for 429 in the response (if available)
                        status_code = getattr(getattr(e, "response", None), "status_code", None)
                        if status_code == 429:
                            delay = rate_limiter.wait_and_retry(attempt)
                            if status_callback:
                                status_callback(f"Rate limited on {metric} row {idx+1}. Waiting {delay:.1f}s...")
                            print(f"429 Too Many Requests: waiting {delay:.1f}s before retry (attempt {attempt+1})")
                        elif status_code == 401:
                            error_msg = "Unauthorized: Check your API key."
                            df.at[idx, f"{metric} Reason"] = error_msg
                            if status_callback:
                                status_callback(error_msg)
                            break  # Do not retry on 401
                        else:
                            error_msg = f"HTTP error ({status_code}) processing {metric} for row {idx+1}: {str(e)}"
                            df.at[idx, f"{metric} Reason"] = error_msg
                            if status_callback:
                                status_callback(error_msg)
                            break
                    except Exception as e:
                        # Handle other exceptions
                        error_msg = f"Error processing {metric} for row {idx+1}: {str(e)}"
                        df.at[idx, f"{metric} Reason"] = error_msg
                        if status_callback:
                            status_callback(error_msg)
                        break
                else:
                    # Max retries exceeded
                    df.at[idx, f"{metric} Reason"] = "Max retries exceeded"

        elapsed_time = time.time() - start_time
        return df, elapsed_time

    def get_row_value(self, row, column_name):
        """Case-insensitive column value retrieval"""
        normalized_columns = {col.lower(): col for col in row.index}
        normalized_name = column_name.lower()
        
        if normalized_name in normalized_columns:
            return row[normalized_columns[normalized_name]]
        return ""

    def save_results(self, df, output_path, add_summary=False):
        """
        Save the processed results to an Excel file, optionally with a summary sheet.
        Applies automatic formatting to the output.
        """
        try:
            if add_summary:
                from batch_processing.summary import add_summary_sheet
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Results')
                    add_summary_sheet(writer, df)
                    self.apply_formatting(writer)
            else:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Results')
                    self.apply_formatting(writer)
            return True
        except Exception as e:
            print(f"Error saving results: {str(e)}")
            return False

    def apply_formatting(self, writer):
        """Apply consistent formatting to Excel output"""
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        # Define styles
        center_alignment = Alignment(horizontal='center', vertical='center')
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # Apply formatting to all columns
        for col_idx, column in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row), 1):
            col_letter = column[0].column_letter
            
            # Set column widths
            worksheet.column_dimensions[col_letter].width = 25
            
            # Apply cell formatting
            for cell in column:
                if cell.row == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.alignment = center_alignment
                else:
                    if "Score" in cell.column_letter or "Status" in cell.column_letter:
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = wrap_alignment
        
        # Auto-adjust row heights for text-heavy columns
        for row in worksheet.iter_rows():
            for cell in row:
                if "Reason" in cell.column_letter or cell.column_letter in ['A', 'B', 'C']:
                    worksheet.row_dimensions[cell.row].height = 60