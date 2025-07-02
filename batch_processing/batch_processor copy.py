import pandas as pd
import importlib
import time
import os
import sys
from pathlib import Path

# Add parent directory to path to import scoring modules
sys.path.append(str(Path(__file__).parent.parent))

class BatchProcessor:
    """
    Handles batch processing of GenAI evaluation tasks from Excel files.
    """

    def __init__(self, api_key, api_url, accept_criteria):
        """
        Initialize the batch processor.

        Parameters:
        - api_key (str): The API key for DeepSeek
        - api_url (str): The API URL for DeepSeek
        - accept_criteria (dict): Dictionary of acceptance thresholds for each metric
        """
        self.api_key = api_key
        self.api_url = api_url
        self.accept_criteria = accept_criteria
        self.metrics = [
            "Correctness", "Relevancy", "Hallucination", "Completeness",
            "Bias", "Toxicity", "Consistency"
        ]
        self.required_columns = [
            "Question to chatbot", "Chatbot Response", "Expected Response"
        ]

    def validate_excel(self, file_path):
        """
        Validates that the Excel file contains the required columns.

        Parameters:
        - file_path (str): Path to the Excel file

        Returns:
        - tuple: (is_valid, error_message)
        """
        try:
            df = pd.read_excel(file_path)
            missing_columns = [col for col in self.required_columns if col not in df.columns]
            if missing_columns:
                return False, f"Missing required columns: {', '.join(missing_columns)}"
            return True, ""
        except Exception as e:
            return False, f"Error validating Excel file: {str(e)}"

    def process_batch(self, file_path, progress_callback=None, stop_flag=None):
        """
        Process all rows in the Excel file for all metrics.

        Parameters:
        - file_path (str): Path to the Excel file
        - progress_callback (function): Optional callback function to report progress
        - stop_flag (callable): Optional function that returns True if processing should stop

        Returns:
        - tuple: (processed_df, elapsed_time)
        """
        start_time = time.time()
        df = pd.read_excel(file_path)
        total_rows = len(df)

        # Create result columns for each metric
        for metric in self.metrics:
            df[f"{metric} Score"] = None
            df[f"{metric} Status"] = None
            df[f"{metric} Reason"] = None

        # Process each row
        for idx, row in df.iterrows():
            if stop_flag and stop_flag():
                break

            question = row.get("Question to chatbot", "")
            chatbot_response = row.get("Chatbot Response", "")
            expected_response = row.get("Expected Response", "")

            # Update progress if callback provided
            if progress_callback:
                progress_callback(idx, total_rows)

            # Process each metric for this row
            for metric in self.metrics:
                if stop_flag and stop_flag():
                    break
                try:
                    module_name = f"scoring_files.{metric.lower()}"
                    scoring_module = importlib.import_module(module_name)
                    if metric == "Correctness":
                        result, _ = scoring_module.evaluate_correctness(
                            question, chatbot_response, expected_response,
                            self.api_key, self.api_url
                        )
                    else:
                        evaluation_function = getattr(scoring_module, f"evaluate_{metric.lower()}")
                        result, _ = evaluation_function(
                            question, chatbot_response,
                            self.api_key, self.api_url
                        )
                    score = result.get("score", 0)
                    threshold = self.accept_criteria.get(metric, 90)
                    # For Bias/Toxicity, you may want to customize pass/fail logic
                    if metric in ["Toxicity", "Bias"]:
                        status = "Failed" if score >= threshold else "Passed"
                    else:
                        status = "Passed" if score >= threshold else "Failed"
                    reason = result.get("reason", "No reason provided")
                    df.at[idx, f"{metric} Score"] = f"{score}%"
                    df.at[idx, f"{metric} Status"] = status
                    df.at[idx, f"{metric} Reason"] = reason
                except Exception as e:
                    df.at[idx, f"{metric} Score"] = "0%"
                    df.at[idx, f"{metric} Status"] = "Error"
                    df.at[idx, f"{metric} Reason"] = f"Error: {str(e)}"

        elapsed_time = time.time() - start_time
        return df, elapsed_time

from openpyxl.styles import Alignment

def save_results(self, df, output_path, add_summary=False):
    """
    Save the processed results to an Excel file, optionally with a summary sheet.
    """
    try:
        if add_summary:
            from batch_processing.summary import add_summary_sheet
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Results')
                add_summary_sheet(writer, df)

                # Center-align all "Score" columns in Results sheet
                workbook = writer.book
                worksheet = writer.sheets['Results']
                for col_idx, column in enumerate(df.columns, 1):
                    if "Score" in column:
                        for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=worksheet.max_row):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Results')
                workbook = writer.book
                worksheet = writer.sheets['Results']
                for col_idx, column in enumerate(df.columns, 1):
                    if "Score" in column:
                        for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=worksheet.max_row):
                            for cell in row:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
        return True
    except Exception as e:
        print(f"Error saving results: {str(e)}")
        return False