import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill

def add_summary_sheet(writer, df):
    """Create summary sheet with pass rates for each metric"""
    metrics = [
        "Correctness", "Relevancy", "Hallucination",
        "Completeness", "Bias", "Toxicity", "Consistency"
    ]
    
    summary_data = []
    for metric in metrics:
        status_col = f"{metric} Status"
        try:
            total = df[status_col].count()
            passed = (df[status_col] == "Passed").sum()
            pass_rate = passed / total if total > 0 else 0
            summary_data.append({
                "Metric": metric,
                "Pass Rate": pass_rate
            })
        except KeyError:
            summary_data.append({
                "Metric": metric,
                "Pass Rate": "N/A"
            })

    # Create DataFrame and write to Excel
    summary_df = pd.DataFrame(summary_data)
    
    # Create new sheet at first position
    summary_df.to_excel(
        writer,
        sheet_name="Summary",
        index=False,
        startrow=2,
        header=False
    )

    # Get workbook objects
    workbook = writer.book
    worksheet = writer.sheets["Summary"]

    # Add title and formatting
    worksheet.merge_cells("A1:B1")
    title_cell = worksheet["A1"]
    title_cell.value = "Evaluation Metrics Summary"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Add column headers
    worksheet["A2"] = "Metric"
    worksheet["B2"] = "Pass Rate"
    for cell in ["A2", "B2"]:
        worksheet[cell].font = Font(bold=True)
        worksheet[cell].alignment = Alignment(horizontal="center")

    # Format data cells
    for row in range(3, len(metrics)+3):
        cell = worksheet[f"B{row}"]
        if isinstance(cell.value, (float, int)):
            # Apply percentage formatting (Excel will multiply by 100)
            cell.number_format = "0.00%"
            cell.fill = PatternFill(
                start_color="F0F8FF",
                end_color="F0F8FF",
                fill_type="solid"
            )
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 18